from django.shortcuts import render, redirect, get_object_or_404
from .models import Grocery, Kitchen, Snack, History
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Q
import csv
from django.http import HttpResponse
import json
from django.contrib import messages
from decimal import Decimal
from django.db.models.functions import TruncMonth
from django.db.models import Sum, Max
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter

# Authentication Views


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
    return render(request, 'accounts/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def home(request):
    return render(request, 'base.html')

# Grocery CRUD

@login_required
def grocery_list(request):
    query = request.GET.get('q')
    if query:
        stocks = Grocery.objects.filter(
            Q(name__icontains=query) |
            Q(remarks__icontains=query)
        )
    else:
        stocks = Grocery.objects.all()
    return render(request, 'groceries/index.html', {'stocks': stocks, 'query': query})

@login_required
def grocery_add(request):
    if request.method == 'POST':
        balance = int(request.POST['balance']) if request.POST['balance'] else 0
        unit_price = float(request.POST['unit_price']) if request.POST['unit_price'] else 0.0
        remarks = request.POST.get('remarks', '')

        stock_in = 0
        stock_out = 0
        total_balance = balance  # since stock_in and stock_out are 0
        total_cost = total_balance * unit_price

        item = Grocery.objects.create(
            name=request.POST['name'],
            balance=balance,
            stock_in = request.POST.get('stock_in', 0),
            stock_out=request.POST.get('stock_out', 0),
            remarks=remarks,
            total_balance=total_balance,
            unit_price=unit_price,
            total_cost=total_cost
        )

        History.objects.create(
            action='Add',
            item_name=item.name,
            category='Grocery',
            user=request.user,
        )
        return redirect('grocery_list')
    return render(request, 'groceries/add.html')



@login_required
def grocery_delete(request, id):
    stock = get_object_or_404(Grocery, id=id)
    if request.method == 'POST':
        History.objects.create(
            action='Delete',
            item_name=stock.name,
            category='Grocery',
            user=request.user
        )
        stock.delete()
        return redirect('grocery_list')
    # Optional: If you're no longer using the delete.html template
    return redirect('grocery_list')

@login_required
def grocery_monthly_summary(request):
    query = request.GET.get('q')

    base_qs = (
        Grocery.objects
        .annotate(month=TruncMonth('date_added'))
        .values('month', 'name', 'unit_price', 'remarks')  # ✅ Include unit_price
        .annotate(
            total_balance=Sum('balance'),
            total_in=Sum('stock_in'),
            total_out=Sum('stock_out'),
            total_cost=Sum('total_cost'),
            latest_remarks=Max('remarks'),
        )
        .order_by('month', 'name')
    )

    if query:
        base_qs = base_qs.filter(name__icontains=query)

    monthly_data = list(base_qs)

    months = sorted(set(entry['month'].strftime('%B %Y') for entry in monthly_data))
    stock_in = []
    stock_out = []
    total_balance = []
    total_cost = []

    month_summary = {
        m: {
            'stock_in': 0,
            'stock_out': 0,
            'total_balance': 0,
            'total_cost': 0,
        } for m in months
    }

    for entry in monthly_data:
        month_label = entry['month'].strftime('%B %Y')
        month_summary[month_label]['stock_in'] += entry['total_in'] or 0
        month_summary[month_label]['stock_out'] += entry['total_out'] or 0
        month_summary[month_label]['total_balance'] += entry['total_balance'] or 0
        month_summary[month_label]['total_cost'] += float(entry['total_cost']) or 0

    for m in months:
        stock_in.append(month_summary[m]['stock_in'])
        stock_out.append(month_summary[m]['stock_out'])
        total_balance.append(month_summary[m]['total_balance'])
        total_cost.append(month_summary[m]['total_cost'])

    context = {
        'monthly_data': monthly_data,
        'months': months,
        'stock_in': stock_in,
        'stock_out': stock_out,
        'total_balance': total_balance,
        'total_cost': total_cost,
    }
    return render(request, 'groceries/monthly_summary.html', context)

@login_required
def kitchen_list(request):
    query = request.GET.get('q')
    if query:
        kitchens = Kitchen.objects.filter(
            Q(name__icontains=query) |
            Q(remarks__icontains=query)
        )
    else:
        kitchens = Kitchen.objects.all()
    return render(request, 'kitchen/index.html', {'items': kitchens, 'query': query})

@login_required
def kitchen_add(request):
    if request.method == 'POST':
        balance = int(request.POST['balance']) if request.POST['balance'] else 0
        unit_price = float(request.POST['unit_price']) if request.POST['unit_price'] else 0.0
        remarks = request.POST.get('remarks', '')

        stock_in = 0
        stock_out = 0
        total_balance = balance  # since stock_in and stock_out are 0
        total_cost = total_balance * unit_price

        item = Kitchen.objects.create(
            name=request.POST['name'],
            balance=balance,
            stock_in=stock_in,
            stock_out=stock_out,
            remarks=remarks,
            total_balance=total_balance,
            unit_price=unit_price,
            total_cost=total_cost
        )

        History.objects.create(
            action='Add',
            item_name=item.name,
            category='Kitchen',
            user=request.user
        )
        return redirect('kitchen_list')
    return render(request, 'kitchen/add.html')





@login_required
def kitchen_delete(request, id):
    stock = get_object_or_404(Kitchen, id=id)
    if request.method == 'POST':
        History.objects.create(
            action='Delete',
            item_name=stock.name,
            category='Kitchen',
            user=request.user
        )
        stock.delete()
    return redirect('kitchen_list')

    
@login_required
def snack_list(request):
    query = request.GET.get('q')
    if query:
        snack_stocks = Snack.objects.filter(
            Q(name__icontains=query) | Q(remarks__icontains=query)
        )
    else:
        snack_stocks = Snack.objects.all()
    return render(request, 'snacks/index.html', {'snack_stocks': snack_stocks, 'query': query})

@login_required
def snack_add(request):
    if request.method == 'POST':
        balance = int(request.POST['balance'])
        stock_in = int(request.POST.get('stock_in', 0))
        stock_out = int(request.POST.get('stock_out', 0))
        unit_price = float(request.POST['unit_price'])
        total = balance + stock_in - stock_out
        total_cost = total * unit_price

        item = Snack.objects.create(
            name=request.POST['name'],
            balance=balance,
            stock_in=stock_in,
            stock_out=stock_out,
            remarks=request.POST.get('remarks', ''),
            total_balance=total,
            unit_price=unit_price,
            total_cost=total_cost
        )

        History.objects.create(
            action='Add',
            item_name=item.name,
            category='Snack',
            user=request.user
        )

        return redirect('snack_list')

    return render(request, 'snacks/add.html')


@login_required
def snack_delete(request, id):
    stock = get_object_or_404(Snack, id=id)
    if request.method == 'POST':
        History.objects.create(
            action='Delete',
            item_name=stock.name,
            category='Snack',
            user=request.user
        )
        stock.delete()
    return redirect('snack_list')

@login_required
def history_list(request):
    history = History.objects.all().order_by('-timestamp')
    return render(request, 'history/index.html', {'history_list': history})


@login_required
def delete_all_history(request):
    if request.method == 'POST':
        History.objects.all().delete()
        messages.success(request, "All history records deleted successfully.")
    return redirect('history_list')

@login_required
def dashboard(request):
    groceries = Grocery.objects.all()
    kitchens = Kitchen.objects.all()
    snacks = Snack.objects.all()

    context = {
        'groceries': groceries,
        'kitchens': kitchens,
        'snacks': snacks,
    }
    return render(request, 'dashboard.html', context)



def export_history_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_history.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date','Time', 'Action', 'Item', 'Category', 'User'])

    history_entries = History.objects.all().order_by('-timestamp')

    for entry in history_entries:
        writer.writerow([
            entry.timestamp.strftime('%Y-%m-%d'),  # Date
            entry.timestamp.strftime('%H:%M'),  # Time
            entry.action,
            entry.item_name,
            entry.category,
            str(entry.user)
        ])

    return response

@login_required
def export_grocery_excel(request):
    query = request.GET.get('q')
    base_qs = (
        Grocery.objects
        .annotate(month=TruncMonth('date_added'))
        .values('month', 'name', 'unit_price', 'remarks')
        .annotate(
            total_balance=Sum('balance'),
            total_in=Sum('stock_in'),
            total_out=Sum('stock_out'),
            total_cost=Sum('total_cost'),
            latest_remarks=Max('remarks'),
        )
        .order_by('month', 'name')
    )
    if query:
        base_qs = base_qs.filter(name__icontains=query)

    data = list(base_qs)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grocery Monthly Summary"

    # Define headers
    headers = ['Stock Name', 'Month', 'Stock In', 'Stock Out', 'Remarks','Unit Price (₱)', 'Total Balance', 'Total Cost (₱)']
    ws.append(headers)

    # Append data rows
    for row in data:
        ws.append([
            row['name'],
            row['month'].strftime('%B %Y') if row['month'] else '',
            row['total_in'] or 0,
            row['total_out'] or 0,
            row['remarks'] or 0,
            float(row['unit_price'] or 0),
            row['total_balance'] or 0,
            float(row['total_cost'] or 0),
        ])

    # Set column widths (optional)
    for i, column_width in enumerate([20, 15, 10, 10, 15, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Prepare HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=grocery_monthly_summary.xlsx'
    wb.save(response)
    return response


@login_required
def export_kitchen_excel(request):
    query = request.GET.get('q')
    base_qs = (
        Kitchen.objects
        .annotate(month=TruncMonth('date_added'))
        .values('month', 'name', 'unit_price', 'remarks')
        .annotate(
            total_balance=Sum('balance'),
            total_in=Sum('stock_in'),
            total_out=Sum('stock_out'),
            total_cost=Sum('total_cost'),
            latest_remarks=Max('remarks'),
        )
        .order_by('month', 'name')
    )
    if query:
        base_qs = base_qs.filter(name__icontains=query)

    data = list(base_qs)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kitchen Monthly Summary"

    # Define headers
    headers = ['Stock Name', 'Month', 'Stock In', 'Stock Out', 'Remarks','Unit Price (₱)', 'Total Balance', 'Total Cost (₱)']
    ws.append(headers)

    # Append data rows
    for row in data:
        ws.append([
            row['name'],
            row['month'].strftime('%B %Y') if row['month'] else '',
            row['total_in'] or 0,
            row['total_out'] or 0,
            row['remarks'] or 0,
            float(row['unit_price'] or 0),
            row['total_balance'] or 0,
            float(row['total_cost'] or 0),
        ])

    # Set column widths (optional)
    for i, column_width in enumerate([20, 15, 10, 10, 15, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Prepare HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=grocery_monthly_summary.xlsx'
    wb.save(response)
    return response



@login_required
def export_snack_excel(request):
    query = request.GET.get('q')
    base_qs = (
        Snack.objects
        .annotate(month=TruncMonth('date_added'))
        .values('month', 'name', 'unit_price', 'remarks')
        .annotate(
            total_balance=Sum('balance'),
            total_in=Sum('stock_in'),
            total_out=Sum('stock_out'),
            total_cost=Sum('total_cost'),
            latest_remarks=Max('remarks'),
        )
        .order_by('month', 'name')
    )
    if query:
        base_qs = base_qs.filter(name__icontains=query)

    data = list(base_qs)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snack Monthly Summary"

    # Define headers
    headers = ['Stock Name', 'Month', 'Stock In', 'Stock Out', 'Remarks','Unit Price (₱)', 'Total Balance', 'Total Cost (₱)']
    ws.append(headers)

    # Append data rows
    for row in data:
        ws.append([
            row['name'],
            row['month'].strftime('%B %Y') if row['month'] else '',
            row['total_in'] or 0,
            row['total_out'] or 0,
            row['remarks'] or 0,
            float(row['unit_price'] or 0),
            row['total_balance'] or 0,
            float(row['total_cost'] or 0),
        ])

    # Set column widths (optional)
    for i, column_width in enumerate([20, 15, 10, 10, 15, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Prepare HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=grocery_monthly_summary.xlsx'
    wb.save(response)
    return response


def stock_in(request, id):
    stock = get_object_or_404(Grocery, id=id)
    
    if request.method == 'POST':
        stock_in_qty = int(request.POST.get('stock_in', 0))
        unit_price_input = request.POST.get('unit_price')

        # Add to balance
        stock.balance += stock_in_qty

        # Try to parse the unit price as Decimal
        try:
            unit_price_decimal = Decimal(unit_price_input)
            if unit_price_decimal >= 0:
                stock.unit_price = unit_price_decimal
        except (TypeError, ValueError, InvalidOperation):
            unit_price_decimal = stock.unit_price  # fallback to current unit_price

        # Compute cost for this batch
        stock_in_cost = Decimal(stock_in_qty) * stock.unit_price

        # Update fields
        stock.stock_in += stock_in_qty
        stock.total_cost += stock_in_cost
        stock.total_balance = stock.balance + stock.stock_in - stock.stock_out

        stock.save()
        return redirect('grocery_list')

    return render(request, 'groceries/stock_in_form.html', {'item': stock})



def stock_out(request, id):
    stock = get_object_or_404(Grocery, id=id)

    if request.method == 'POST':
        stock_out_qty = int(request.POST.get('stock_out', 0))
        
        # Prevent removing more than available
        stock_out_qty = min(stock.balance, stock_out_qty)

        # Calculate cost of items being removed
        stock_out_cost = Decimal(stock_out_qty) * stock.unit_price

        # Update fields
        stock.stock_out += stock_out_qty
        stock.balance -= stock_out_qty                    # <-- Subtract from balance
        stock.total_balance = stock.balance               # <-- Update total_balance to match
        stock.total_cost -= stock_out_cost                # <-- Subtract cost

        # Avoid negative cost
        if stock.total_cost < 0:
            stock.total_cost = Decimal('0.00')

        stock.save()
        return redirect('grocery_list')

    return render(request, 'groceries/stock_out_form.html', {'item': stock})


def kitchen_stock_in(request, id):
    stock = get_object_or_404(Kitchen, id=id)
    
    if request.method == 'POST':
        stock_in_qty = int(request.POST.get('stock_in', 0))
        unit_price_input = request.POST.get('unit_price')

        # Add to balance
        stock.balance += stock_in_qty

        # Try to parse the unit price as Decimal
        try:
            unit_price_decimal = Decimal(unit_price_input)
            if unit_price_decimal >= 0:
                stock.unit_price = unit_price_decimal
        except (TypeError, ValueError, InvalidOperation):
            unit_price_decimal = stock.unit_price  # fallback to current unit_price

        # Compute cost for this batch
        stock_in_cost = Decimal(stock_in_qty) * stock.unit_price

        # Update fields
        stock.stock_in += stock_in_qty
        stock.total_cost += stock_in_cost
        stock.total_balance = stock.balance + stock.stock_in - stock.stock_out

        stock.save()
        return redirect('kitchen_list')  # Redirect to kitchen list

    return render(request, 'kitchen/stock_in_form.html', {'item': stock})



def kitchen_stock_out(request, id):
    stock = get_object_or_404(Kitchen, id=id)

    if request.method == 'POST':
        stock_out_qty = int(request.POST.get('stock_out', 0))

        # Prevent subtracting more than available
        stock_out_qty = min(stock.total_balance, stock_out_qty)

        # Calculate cost of stock going out
        stock_out_cost = Decimal(stock_out_qty) * stock.unit_price

        # Update stock fields
        stock.stock_out += stock_out_qty
        stock.balance -= stock_out_qty  # <-- this line ensures balance is updated
        stock.total_balance -= stock_out_qty
        stock.total_cost -= stock_out_cost

        # Prevent negative values
        if stock.total_cost < 0:
            stock.total_cost = Decimal('0.00')
        if stock.balance < 0:
            stock.balance = 0
        if stock.total_balance < 0:
            stock.total_balance = 0

        stock.save()
        return redirect('kitchen_list')

    return render(request, 'kitchen/stock_out_form.html', {'item': stock})


@login_required
def snack_stock_in(request, id):
    stock = get_object_or_404(Snack, id=id)
    
    if request.method == 'POST':
        stock_in_qty = int(request.POST.get('stock_in', 0))
        unit_price_input = request.POST.get('unit_price')

        # Add to balance
        stock.balance += stock_in_qty

        # Try to parse the unit price as Decimal
        try:
            unit_price_decimal = Decimal(unit_price_input)
            if unit_price_decimal >= 0:
                stock.unit_price = unit_price_decimal
        except (TypeError, ValueError, InvalidOperation):
            unit_price_decimal = stock.unit_price  # fallback to current unit_price

        # Compute cost for this batch
        stock_in_cost = Decimal(stock_in_qty) * stock.unit_price

        # Update fields
        stock.stock_in += stock_in_qty
        stock.total_cost += stock_in_cost
        stock.total_balance = stock.balance + stock.stock_in - stock.stock_out

        stock.save()
        return redirect('snack_list')

    return render(request, 'snacks/stock_in_form.html', {'item': stock})
@login_required
def snack_stock_out(request, id):
    stock = get_object_or_404(Snack, id=id)

    if request.method == 'POST':
        stock_out_qty = int(request.POST.get('stock_out', 0))
        
        # Prevent removing more than available balance
        stock_out_qty = min(stock.balance, stock_out_qty)

        # Calculate cost of items being removed
        stock_out_cost = Decimal(stock_out_qty) * stock.unit_price

        # Update fields
        stock.stock_out += stock_out_qty
        stock.balance -= stock_out_qty                    # Subtract from balance
        stock.total_balance = stock.balance               # Update total_balance to match
        stock.total_cost -= stock_out_cost                # Subtract cost

        # Avoid negative cost
        if stock.total_cost < 0:
            stock.total_cost = Decimal('0.00')

        stock.save()
        return redirect('snack_list')

    return render(request, 'snacks/stock_out_form.html', {'item': stock})

@login_required
def kitchen_monthly_summary(request):
    query = request.GET.get('q')

    base_qs = (
        Kitchen.objects
        .annotate(month=TruncMonth('date_added'))
        .values('month', 'name', 'unit_price', 'remarks')  # ✅ Include unit_price
        .annotate(
            total_balance=Sum('balance'),
            total_in=Sum('stock_in'),
            total_out=Sum('stock_out'),
            total_cost=Sum('total_cost'),
            latest_remarks=Max('remarks'),
        )
        .order_by('month', 'name')
    )

    if query:
        base_qs = base_qs.filter(name__icontains=query)

    monthly_data = list(base_qs)

    months = sorted(set(entry['month'].strftime('%B %Y') for entry in monthly_data))
    stock_in = []
    stock_out = []
    total_balance = []
    total_cost = []

    month_summary = {
        m: {
            'stock_in': 0,
            'stock_out': 0,
            'total_balance': 0,
            'total_cost': 0
        } for m in months
    }

    for entry in monthly_data:
        month_label = entry['month'].strftime('%B %Y')
        month_summary[month_label]['stock_in'] += entry['total_in'] or 0
        month_summary[month_label]['stock_out'] += entry['total_out'] or 0
        month_summary[month_label]['total_balance'] += entry['total_balance'] or 0
        month_summary[month_label]['total_cost'] += float(entry['total_cost']) or 0

    for m in months:
        stock_in.append(month_summary[m]['stock_in'])
        stock_out.append(month_summary[m]['stock_out'])
        total_balance.append(month_summary[m]['total_balance'])
        total_cost.append(month_summary[m]['total_cost'])

    context = {
        'monthly_data': monthly_data,
        'months': months,
        'stock_in': stock_in,
        'stock_out': stock_out,
        'total_balance': total_balance,
        'total_cost': total_cost,
    }
    return render(request, 'kitchen/monthly_summary.html', context)


@login_required
def snack_monthly_summary(request):
    query = request.GET.get('q')

    base_qs = (
        Snack.objects
        .annotate(month=TruncMonth('date_added'))
        .values('month', 'name', 'unit_price', 'remarks')  # ✅ Include unit_price
        .annotate(
            total_balance=Sum('balance'),
            total_in=Sum('stock_in'),
            total_out=Sum('stock_out'),
            total_cost=Sum('total_cost'),
            latest_remarks=Max('remarks'),
        )
        .order_by('month', 'name')
    )

    if query:
        base_qs = base_qs.filter(name__icontains=query)

    monthly_data = list(base_qs)

    months = sorted(set(entry['month'].strftime('%B %Y') for entry in monthly_data))
    stock_in = []
    stock_out = []
    total_balance = []
    total_cost = []

    month_summary = {
        m: {
            'stock_in': 0,
            'stock_out': 0,
            'total_balance': 0,
            'total_cost': 0
        } for m in months
    }

    for entry in monthly_data:
        month_label = entry['month'].strftime('%B %Y')
        month_summary[month_label]['stock_in'] += entry['total_in'] or 0
        month_summary[month_label]['stock_out'] += entry['total_out'] or 0
        month_summary[month_label]['total_balance'] += entry['total_balance'] or 0
        month_summary[month_label]['total_cost'] += float(entry['total_cost']) or 0

    for m in months:
        stock_in.append(month_summary[m]['stock_in'])
        stock_out.append(month_summary[m]['stock_out'])
        total_balance.append(month_summary[m]['total_balance'])
        total_cost.append(month_summary[m]['total_cost'])

    context = {
        'monthly_data': monthly_data,
        'months': months,
        'stock_in': stock_in,
        'stock_out': stock_out,
        'total_balance': total_balance,
        'total_cost': total_cost,
    }
    return render(request, 'snacks/monthly_summary.html', context)



